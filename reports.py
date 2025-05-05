# reports.py
import os
import logging
from pathlib import Path
from datetime import datetime

from reportlab.lib import colors
from reportlab.lib.pagesizes import LETTER, landscape
from reportlab.lib.units import mm
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont

from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment

logger = logging.getLogger(__name__)

# --- Настройки страницы ---
PAGE_SIZE = LETTER
MARGIN    = 15 * mm

# --- Поиск системного TTF-шрифта ---
FONT_NAME   = None
font_paths = []

import sys
if sys.platform.startswith("win"):
    windir = Path(os.environ.get("WINDIR", "C:/Windows"))
    font_paths += [
        windir / "Fonts" / "arial.ttf",
        windir / "Fonts" / "ARIAL.TTF",
    ]
elif sys.platform == "darwin":
    font_paths += [
        Path("/Library/Fonts/Arial.ttf"),
        Path("/Library/Fonts/DejaVuSans.ttf"),
    ]
else:
    # Linux
    font_paths += [
        Path("/usr/share/fonts/truetype/msttcorefonts/Arial.ttf"),
        Path("/usr/share/fonts/truetype/msttcorefonts/arial.ttf"),
        Path("/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf"),
        Path("/usr/share/fonts/truetype/freefont/FreeSans.ttf"),
    ]

# Регистрируем первый из найденных
for p in font_paths:
    if p.exists():
        try:
            name = p.stem  # например "arial" или "DejaVuSans"
            pdfmetrics.registerFont(TTFont(name, str(p)))
            FONT_NAME = name
            logger.info(f"Используется системный шрифт: {name} ({p})")
            break
        except Exception as e:
            logger.warning(f"Ошибка регистрации {p}: {e}")

# Если ничего не нашли — fallback на Helvetica
if not FONT_NAME:
    FONT_NAME = "Helvetica"
    logger.warning("Не найден system TTF, используется Helvetica (кириллица может не отображаться)")

# --- Стили Platypus ---
styles = getSampleStyleSheet()
styles["Normal"].fontName    = FONT_NAME
styles["Normal"].fontSize    = 10
styles["Heading1"].fontName  = FONT_NAME
styles["Heading1"].fontSize  = 14

def export_employees_pdf(path: str, employees):
    """Экспорт сотрудников в PDF с system-шрифтом."""
    doc = SimpleDocTemplate(path,
        pagesize=landscape(PAGE_SIZE),
        leftMargin=MARGIN, rightMargin=MARGIN,
        topMargin=MARGIN, bottomMargin=MARGIN
    )
    elems = []
    elems.append(Paragraph(f"Отчёт сотрудников — {datetime.now():%Y-%m-%d}", styles["Heading1"]))
    elems.append(Spacer(1, 6*mm))

    data = [["ID","ФИО","Должность","Дата приёма","Отпуск (дн.)"]]
    for e in employees:
        data.append([
            str(e.id),
            f"{e.first_name} {e.last_name}",
            e.position or "",
            e.hire_date.strftime("%d.%m.%Y") if e.hire_date else "",
            str(e.vacation_days_left)
        ])

    table = Table(data, colWidths=[20*mm,60*mm,50*mm,40*mm,30*mm], repeatRows=1)
    table.setStyle(TableStyle([
        ("FONTNAME",(0,0),(-1,-1),FONT_NAME),
        ("FONTSIZE",(0,0),(-1,-1),10),
        ("ALIGN",(0,0),(-1,0),"CENTER"),
        ("VALIGN",(0,0),(-1,-1),"MIDDLE"),
        ("GRID",(0,0),(-1,-1),0.5,colors.gray),
        ("BACKGROUND",(0,0),(-1,0),colors.lightgrey),
    ]))
    elems.append(table)
    doc.build(elems)

def export_employees_excel(path: str, employees):
    """Экспорт сотрудников в Excel."""
    wb = Workbook()
    ws = wb.active; ws.title = "Сотрудники"
    headers = ["ID","ФИО","Должность","Дата приёма","Отпуск (дн.)"]
    ws.append(headers)
    for e in employees:
        ws.append([
            e.id,
            f"{e.first_name} {e.last_name}",
            e.position or "",
            e.hire_date.strftime("%d.%m.%Y") if e.hire_date else "",
            e.vacation_days_left
        ])
    for col in ws.columns:
        max_len = max(len(str(c.value)) for c in col if c.value is not None)
        letter = get_column_letter(col[0].column)
        ws.column_dimensions[letter].width = max_len + 2
        for cell in col:
            cell.alignment = Alignment(horizontal="left", vertical="center")
    wb.save(path)
